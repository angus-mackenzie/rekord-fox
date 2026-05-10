"""Tracklist exporters: TXT (default), CSV, XLSX.

Each builder takes the full timeline view (segments + manual tags) and
returns `(bytes, content_type, filename)`. Endpoints just dispatch on the
URL extension and stream the result.

Manual tags ride alongside auto-detected segments and are flagged via the
`kind` column in CSV/XLSX. TXT is intentionally minimal — just title,
artist, and one link per row — since the user's stated use case is
"hand to a friend / paste into a doc".
"""
from __future__ import annotations

import csv
import io
from collections.abc import Iterable
from datetime import datetime
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

Format = Literal["txt", "csv", "xlsx"]


def fmt_time(s: float) -> str:
    s = max(0, int(s))
    h, rem = divmod(s, 3600)
    m, sec = divmod(rem, 60)
    return f"{h}:{m:02d}:{sec:02d}" if h else f"{m}:{sec:02d}"


def fmt_dur(s: float) -> str:
    s = int(s)
    m, sec = divmod(s, 60)
    return f"{m}m {sec}s" if m else f"{sec}s"


# Order in which we pick the "single best" link to surface in TXT output.
# Spotify is the most useful for handing a tracklist to someone else;
# Shazam is the canonical source-of-truth fallback.
TXT_LINK_PRIORITY = ("spotify", "shazam", "youtube", "soundcloud", "apple_music")


def _pick_txt_link(urls: dict[str, str] | None) -> str | None:
    if not urls:
        return None
    for k in TXT_LINK_PRIORITY:
        v = urls.get(k)
        if v and v.startswith(("http://", "https://")):
            return v
    return None


def _row_dicts(
    segments: Iterable[dict[str, Any]],
    manual_tags: Iterable[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Merge segments + manual tags into a unified, time-sorted list of rows.

    Each row has a `kind` ('segment' | 'manual') and a normalised set of
    fields the exporters can render uniformly. Used by all three formats so
    they always agree on row order and content.
    """
    rows: list[dict[str, Any]] = []
    for s in segments:
        primary = (s.get("candidates") or [{}])[0] if s.get("candidates") else {}
        rows.append({
            "kind": "segment",
            "start": s["start_seconds"],
            "end": s["end_seconds"],
            "title": s.get("title"),
            "artist": s.get("artist"),
            "state": s.get("state"),
            "confidence": s.get("confidence") or 0.0,
            "external_urls": primary.get("external_urls") or {},
            "notes": s.get("notes"),
        })
    for t in manual_tags:
        rows.append({
            "kind": "manual",
            "start": t["start_seconds"],
            "end": t["end_seconds"],
            "title": t.get("title"),
            "artist": t.get("artist"),
            "state": "manual",
            "confidence": 1.0,
            "external_urls": t.get("external_urls") or {},
            "notes": t.get("notes"),
        })
    rows.sort(key=lambda r: (r["start"], r["kind"]))
    return rows


def _slug(name: str) -> str:
    """Filename-safe rendering of the source media name. ASCII-only, dashes."""
    keep = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._-"
    cleaned = "".join(c if c in keep else "-" for c in name)
    # Collapse runs of dashes; trim leading/trailing.
    while "--" in cleaned:
        cleaned = cleaned.replace("--", "-")
    return cleaned.strip("-") or "tracklist"


def _filename(media_filename: str, ext: Format) -> str:
    base = _slug(media_filename.rsplit(".", 1)[0])
    return f"{base}.tracklist.{ext}"


# --- builders -------------------------------------------------------------


def build_txt(
    media_filename: str,
    duration_seconds: float,
    segments: Iterable[dict[str, Any]],
    manual_tags: Iterable[dict[str, Any]],
) -> tuple[bytes, str, str]:
    """Plain-text tracklist — minimal, hand-readable, copy-pasteable.

    Skips unresolved/unidentified gaps to keep the output focused on
    actual tracks. State is annotated only when not 'confirmed' so the
    common case stays clean.
    """
    rows = [
        r for r in _row_dicts(segments, manual_tags)
        if r["title"] or r["artist"]
    ]

    out = io.StringIO()
    out.write("Rekord Fox tracklist\n")
    out.write("=" * 20 + "\n\n")
    out.write(f"Mix: {media_filename}\n")
    out.write(f"Duration: {fmt_time(duration_seconds)}\n")
    out.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")

    for r in rows:
        time_range = f"{fmt_time(r['start'])} – {fmt_time(r['end'])}"
        title = r["title"] or "(untitled)"
        artist = r["artist"] or ""
        title_line = f"  {time_range:>13}   {title}"
        if artist:
            title_line += f" — {artist}"
        # Annotate only weaker matches + manual tags so confident rows stay clean.
        if r["state"] not in (None, "confirmed"):
            title_line += f"  [{r['state']}]"
        out.write(title_line + "\n")
        link = _pick_txt_link(r["external_urls"])
        if link:
            # Indent under the title for visual grouping, matches column.
            out.write(" " * 18 + link + "\n")
        out.write("\n")

    data = out.getvalue().encode("utf-8")
    return data, "text/plain; charset=utf-8", _filename(media_filename, "txt")


# CSV column order — chosen so the most-useful columns are leftmost when
# opened in a spreadsheet without resizing.
CSV_HEADERS = [
    "start", "end", "duration",
    "title", "artist", "state", "confidence",
    "kind", "spotify", "shazam", "youtube", "soundcloud", "apple_music", "notes",
]


def _csv_row(r: dict[str, Any]) -> list[Any]:
    urls = r["external_urls"] or {}
    return [
        fmt_time(r["start"]),
        fmt_time(r["end"]),
        fmt_dur(r["end"] - r["start"]),
        r["title"] or "",
        r["artist"] or "",
        r["state"] or "",
        round(r["confidence"], 3),
        r["kind"],
        urls.get("spotify", ""),
        urls.get("shazam", ""),
        urls.get("youtube", ""),
        urls.get("soundcloud", ""),
        urls.get("apple_music", ""),
        r["notes"] or "",
    ]


def build_csv(
    media_filename: str,
    duration_seconds: float,
    segments: Iterable[dict[str, Any]],
    manual_tags: Iterable[dict[str, Any]],
) -> tuple[bytes, str, str]:
    """CSV with every row including unresolved gaps — meant for processing,
    not for human reading."""
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(CSV_HEADERS)
    for r in _row_dicts(segments, manual_tags):
        w.writerow(_csv_row(r))
    data = out.getvalue().encode("utf-8")
    return data, "text/csv; charset=utf-8", _filename(media_filename, "csv")


def build_xlsx(
    media_filename: str,
    duration_seconds: float,
    segments: Iterable[dict[str, Any]],
    manual_tags: Iterable[dict[str, Any]],
) -> tuple[bytes, str, str]:
    """XLSX with formatted header row + auto-sized columns + clickable
    hyperlinks. Same data as CSV, just nicer to open."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("openpyxl returned no active sheet")
    ws.title = "Tracklist"

    # Title block above the table — context for whoever opens this file.
    ws["A1"] = "Rekord Fox tracklist"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Mix: {media_filename}"
    ws["A3"] = f"Duration: {fmt_time(duration_seconds)}"
    ws["A4"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    header_row = 6
    header_fill = PatternFill("solid", fgColor="22272E")
    header_font = Font(bold=True, color="E7E5EA")
    for i, h in enumerate(CSV_HEADERS, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left")

    # Track maximum width per column for auto-sizing at the end.
    col_widths = [len(h) for h in CSV_HEADERS]
    url_columns = {
        CSV_HEADERS.index("spotify") + 1,
        CSV_HEADERS.index("shazam") + 1,
        CSV_HEADERS.index("youtube") + 1,
        CSV_HEADERS.index("soundcloud") + 1,
        CSV_HEADERS.index("apple_music") + 1,
    }

    rows = _row_dicts(segments, manual_tags)
    for r_offset, r in enumerate(rows, start=header_row + 1):
        values = _csv_row(r)
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=r_offset, column=i, value=v)
            if i in url_columns and isinstance(v, str) and v.startswith(("http://", "https://")):
                cell.hyperlink = v
                cell.font = Font(color="3B82F6", underline="single")
            col_widths[i - 1] = max(col_widths[i - 1], len(str(v)))

    # Cap auto width so very long URLs / album titles don't blow out the sheet.
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(60, w + 2)
    # Freeze the header so it stays visible while scrolling.
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return (
        buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        _filename(media_filename, "xlsx"),
    )


BUILDERS = {
    "txt": build_txt,
    "csv": build_csv,
    "xlsx": build_xlsx,
}
