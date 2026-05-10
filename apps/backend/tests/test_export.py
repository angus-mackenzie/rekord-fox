"""Tracklist export builders + endpoint dispatch."""
from __future__ import annotations

import csv
import importlib
import io
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook


# --- direct builder tests (no DB / HTTP layer) ----------------------------


SAMPLE_SEGMENTS: list[dict[str, Any]] = [
    {
        "start_seconds": 0.0, "end_seconds": 95.0,
        "title": "Pacific", "artist": "Billy Bahama",
        "state": "confirmed", "confidence": 0.9,
        "candidates": [{"external_urls": {"spotify": "https://open.spotify.com/track/abc"}}],
        "notes": None,
    },
    {
        "start_seconds": 90.0, "end_seconds": 155.0,
        "title": None, "artist": None,
        "state": "unresolved", "confidence": 0.0,
        "candidates": [], "notes": "no provider match",
    },
    {
        "start_seconds": 150.0, "end_seconds": 275.0,
        "title": "Puro Urbano (Continuous Mix)", "artist": "Sebas Ramis",
        "state": "uncertain", "confidence": 0.9,
        "candidates": [{"external_urls": {"shazam": "https://shazam.com/track/123"}}],
        "notes": None,
    },
]
SAMPLE_TAGS: list[dict[str, Any]] = [
    {
        "start_seconds": 300.0, "end_seconds": 360.0,
        "title": "Hand-tagged Track", "artist": "Mystery DJ",
        "external_urls": {"spotify": "https://open.spotify.com/track/manual"},
        "notes": "killer transition here",
    },
]


def test_txt_skips_unresolved_and_includes_links():
    from rekord.api.export import build_txt

    data, ctype, name = build_txt("Test Mix.mp3", 600.0, SAMPLE_SEGMENTS, SAMPLE_TAGS)
    text = data.decode()
    assert ctype.startswith("text/plain")
    assert name.endswith(".tracklist.txt")
    # Header info
    assert "Rekord Fox tracklist" in text
    assert "Test Mix.mp3" in text
    # Confirmed track + spotify link present
    assert "Pacific" in text
    assert "Billy Bahama" in text
    assert "open.spotify.com/track/abc" in text
    # Unresolved gap is skipped (no "no provider match" text)
    assert "no provider match" not in text
    # Uncertain state is annotated, but not for the confident one
    assert "[uncertain]" in text
    assert "[confirmed]" not in text
    # Manual tag included
    assert "Hand-tagged Track" in text


def test_txt_picks_spotify_over_shazam_when_both_present():
    from rekord.api.export import build_txt

    seg = dict(SAMPLE_SEGMENTS[0])
    seg["candidates"] = [{
        "external_urls": {
            "spotify": "https://open.spotify.com/spotify-link",
            "shazam": "https://shazam.com/shazam-link",
        }
    }]
    text = build_txt("x.mp3", 60, [seg], [])[0].decode()
    assert "spotify-link" in text
    assert "shazam-link" not in text


def test_csv_includes_every_row_and_correct_headers():
    from rekord.api.export import build_csv

    data, ctype, name = build_csv("Test.mp3", 600.0, SAMPLE_SEGMENTS, SAMPLE_TAGS)
    assert ctype.startswith("text/csv")
    assert name.endswith(".tracklist.csv")
    rows = list(csv.reader(io.StringIO(data.decode())))
    headers = rows[0]
    assert headers[:4] == ["start", "end", "duration", "title"]
    assert "spotify" in headers
    assert "kind" in headers
    # 3 segments + 1 manual tag = 4 data rows. The unresolved one is INCLUDED
    # in csv (unlike txt) since this format is meant for processing.
    assert len(rows) - 1 == 4
    # Manual tag has kind='manual' and points at its spotify link
    manual_row = next(r for r in rows[1:] if r[headers.index("kind")] == "manual")
    assert manual_row[headers.index("title")] == "Hand-tagged Track"
    assert "spotify.com/track/manual" in manual_row[headers.index("spotify")]


def test_xlsx_is_valid_workbook_with_hyperlinks():
    from rekord.api.export import build_xlsx

    data, ctype, name = build_xlsx("Test Mix.mp3", 600.0, SAMPLE_SEGMENTS, SAMPLE_TAGS)
    assert ctype.startswith("application/vnd.openxml")
    assert name.endswith(".tracklist.xlsx")
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    # Title block + header row + 4 data rows
    assert ws["A1"].value == "Rekord Fox tracklist"
    # Find the header row by scanning
    header_row = None
    for r in range(1, 12):
        if ws.cell(row=r, column=1).value == "start":
            header_row = r
            break
    assert header_row is not None
    # Spotify URL in the data should have been turned into a clickable hyperlink
    spotify_col = next(
        i for i in range(1, 20)
        if ws.cell(row=header_row, column=i).value == "spotify"
    )
    pacific_row = header_row + 1
    cell = ws.cell(row=pacific_row, column=spotify_col)
    assert cell.value == "https://open.spotify.com/track/abc"
    assert cell.hyperlink is not None


def test_filename_slugifies_problematic_chars():
    from rekord.api.export import build_txt

    _, _, name = build_txt("Ed/burg @ 404 — \"set\".mp3", 60, [], [])
    # No path separators, quotes, or spaces in the slug
    assert "/" not in name and '"' not in name and " " not in name
    assert name.endswith(".tracklist.txt")


# --- endpoint dispatch ----------------------------------------------------


@pytest.fixture()
def client(monkeypatch, tmp_path):
    monkeypatch.setenv("REKORD_DB_PATH", str(tmp_path / "test.db"))
    monkeypatch.setenv("REKORD_UPLOAD_DIR", str(tmp_path / "uploads"))
    monkeypatch.setenv("REKORD_DATA_DIR", str(tmp_path))
    monkeypatch.setenv("REKORD_PROVIDER", "fake_csv")
    import rekord.config as cfg
    import rekord.db as db
    importlib.reload(cfg)
    importlib.reload(db)
    import rekord.api.main as api
    importlib.reload(api)
    with TestClient(api.app) as c:
        yield c, db


def _seed(db, tmp_path):
    """Job + media + one segment + one manual tag, enough to round-trip
    the export endpoint for each format."""
    from rekord.models import (
        AnalysisJob,
        JobStatus,
        ManualCorrection,
        MediaAsset,
        SegmentState,
        TimelineSegment,
    )

    audio = tmp_path / "x.mp3"
    audio.write_bytes(b"\x00")
    with db.session_scope() as s:
        media = MediaAsset(
            original_filename="Sample Mix.mp3",
            content_type="audio/mp3",
            storage_path=str(audio),
            checksum="0",
            duration_seconds=600.0,
        )
        s.add(media)
        s.commit()
        job = AnalysisJob(media_asset_id=media.id, status=JobStatus.succeeded, progress=1.0)
        s.add(job)
        s.commit()
        s.add(TimelineSegment(
            analysis_job_id=job.id, start_seconds=0.0, end_seconds=60.0,
            state=SegmentState.confirmed, confidence=0.9,
            title="Pacific", artist="Billy Bahama",
            candidates=[{"external_urls": {"spotify": "https://x"}}],
        ))
        s.add(ManualCorrection(
            analysis_job_id=job.id, start_seconds=120.0, end_seconds=180.0,
            title="Manual", artist="Tester",
            external_urls={"shazam": "https://y"},
        ))
        s.commit()
        return job.id


def test_endpoint_dispatches_each_format(client, tmp_path):
    c, db = client
    jid = _seed(db, tmp_path)
    for ext, expected_ctype in [
        ("txt", "text/plain"),
        ("csv", "text/csv"),
        ("xlsx", "application/vnd.openxml"),
    ]:
        r = c.get(f"/jobs/{jid}/tracklist.{ext}")
        assert r.status_code == 200, (ext, r.text)
        assert r.headers["content-type"].startswith(expected_ctype)
        # Filename hint includes the slugged mix name + the requested ext
        cd = r.headers.get("content-disposition", "")
        assert f".tracklist.{ext}" in cd
        assert "Sample-Mix" in cd  # slugged from "Sample Mix.mp3"


def test_endpoint_400_on_unknown_format(client, tmp_path):
    c, db = client
    jid = _seed(db, tmp_path)
    r = c.get(f"/jobs/{jid}/tracklist.json")
    assert r.status_code == 400


def test_endpoint_404_on_unknown_job(client):
    c, _ = client
    r = c.get("/jobs/does-not-exist/tracklist.txt")
    assert r.status_code == 404
